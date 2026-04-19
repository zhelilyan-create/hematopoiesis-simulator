"""Update calibration workbook with 27 real sweep results."""
import json, sys
from pathlib import Path
import openpyxl

WB_PATH = Path('results/calibtations/hemotopoiesis_calibration.xlsx')

# ── load sweep results ──────────────────────────────────────────────────────
raw = open('scripts/_sweep_out.txt').read()
json_block = raw.split('JSON_START\n')[1].split('\nJSON_END')[0].strip()
sweep = json.loads(json_block)

# ── baseline params (fixed, from baseline YAML + not swept) ─────────────────
CENT_STEM    = 0.015
AGE_CAP      = 10
STRESS_ACC   = 0.001
STEM_DRIFT   = 0.0
W_DIV_STEM   = 1.0
W_DIV_REPL   = 0.005
W_DIFF_STRESS= 0.1
W_DIFF_STEM  = 1.0
W_DIFF_REPL  = 0.005
W_APO_STRESS = 1.0
W_APO_REPL   = 0.005

wb = openpyxl.load_workbook(WB_PATH)
ws = wb.active

# ── read header row to get column index map ──────────────────────────────────
header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
col = {h: i+1 for i, h in enumerate(header_row) if h is not None}
print("Columns found:", list(col.keys()))

# ── find next empty data row ─────────────────────────────────────────────────
# scan from bottom for first non-empty run_id row
insert_row = ws.max_row + 1
for r in range(ws.max_row, 1, -1):
    v = ws.cell(r, col['run_id']).value
    if v is not None:
        insert_row = r + 1
        break
print(f"Inserting from row {insert_row}")

# ── add a section header row ─────────────────────────────────────────────────
ws.cell(insert_row, col['run_id']).value = \
    'SERIES 10 — Local sensitivity sweep: centriole_stress_factor × w_div_stress (3×3×3 = 27 runs)'
insert_row += 1

# ── write 27 rows ────────────────────────────────────────────────────────────
def set_flag(r):
    if r['csf'] == 0.01 and r['wds'] == 0.25 and r['seed'] == 42:
        return 'current_best'
    if r['status'] == 'too_stem_heavy':
        return 'rejected'
    return 'candidate'

def hypothesis(r):
    return f"local_sensitivity: csf={r['csf']}, w_div_stress={r['wds']:.2f}, seed={r['seed']}"

for r in sweep:
    row = insert_row
    def w(key, val):
        if key in col:
            ws.cell(row, col[key]).value = val

    w('run_id',          r['run_id'])
    w('mode',            'centriole')
    w('seed',            r['seed'])
    w('t_max',           100.0)
    w('init_hsc',        10)
    w('cent_stem',       CENT_STEM)
    w('cent_stress',     r['csf'])
    w('age_cap',         AGE_CAP)
    w('stress_acc',      STRESS_ACC)
    w('stem_drift',      STEM_DRIFT)
    w('w_div_stress',    r['wds'])
    w('w_div_stemness',  W_DIV_STEM)
    w('w_div_repl',      W_DIV_REPL)
    w('w_diff_stress',   W_DIFF_STRESS)
    w('w_diff_stemness', W_DIFF_STEM)
    w('w_diff_repl',     W_DIFF_REPL)
    w('w_apo_stress',    W_APO_STRESS)
    w('w_apo_repl',      W_APO_REPL)
    w('final_n',         r['final_n'])
    w('HSC',             r['HSC'])
    w('HSC_%',           r['HSC_pct'])
    w('MPP',             r['MPP'])
    w('MPP_%',           r['MPP_pct'])
    w('events',          r['n_events'])
    w('notes',           r['status'])
    w('series',          'local_sensitivity_sweep')
    w('CLP_%',           r['CLP_pct'])
    w('CMP_%',           r['CMP_pct'])
    w('B_cell_%',        r['B_cell_pct'])
    w('T_cell_%',        r['T_cell_pct'])
    w('Myeloid_%',       r['Myeloid_pct'])
    w('Erythroid_%',     r['Erythroid_pct'])
    w('mature_%',        r['mature_pct'])
    w('progenitor_%',    r['progenitor_pct'])
    w('HSC_to_MPP_ratio', r['hsc_mpp_ratio'])
    w('baseline_flag',   set_flag(r))
    w('hypothesis',      hypothesis(r))

    insert_row += 1

wb.save(WB_PATH)
print(f"Saved. Total rows now: {ws.max_row}")

# ── verify current_best count ────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(WB_PATH)
ws2 = wb2.active
hdr2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
bf_col = hdr2.index('baseline_flag') + 1
flags = {}
for r in range(2, ws2.max_row + 1):
    v = ws2.cell(r, bf_col).value
    if v:
        flags[v] = flags.get(v, 0) + 1
print("baseline_flag summary:", flags)
