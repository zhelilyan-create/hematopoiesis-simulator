"""
Append 12 lineage-epigenetic test rows (run_ids 168-179) to calibration workbook.
Series 16 — w_lineage_epigenetic in {0.0, 0.03, 0.05, 0.07}, seeds {42, 1, 99}.
w_diff_epigenetic fixed at 0.0 throughout.
"""
import json, re
from collections import Counter
import openpyxl

WORKBOOK_PATH = "results/calibtations/hemotopoiesis_calibration.xlsx"
SWEEP_OUT     = "scripts/_sweep7_out.txt"

FIXED = dict(
    mode_inh="centriole",
    t_max=100.0,
    init_hsc=10,
    cent_stem=0.015,
    cent_stress=0.01,
    age_cap=10,
    stress_acc=0.001,
    stem_drift=0.0,
    w_div_stemness=1.0,
    w_div_stress=0.25,
    w_div_repl=0.005,
    w_diff_stress=0.1,
    w_diff_stemness=1.0,
    w_diff_repl=0.005,
    w_apo_stress=1.0,
    w_apo_stemness=0.0,
    w_apo_repl=0.005,
    w_diff_epigenetic=0.0,
    series="v0.7_lineage_epigenetic_test",
    hypothesis="epigenetic_bias redistributes lineage commitment (myeloid vs lymphoid)",
)

with open(SWEEP_OUT) as f:
    text = f.read()

m = re.search(r"JSON_START\n(\[.*?\])\nJSON_END", text, re.DOTALL)
runs = json.loads(m.group(1))

wb = openpyxl.load_workbook(WORKBOOK_PATH)
ws = wb.active

header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
col_idx = {h: i + 1 for i, h in enumerate(header) if h}

print(f"Current last row: {ws.max_row}")

insert_row = ws.max_row + 1
ws.cell(insert_row, 1).value = (
    "SERIES 16 — v0.7 lineage-epigenetic test, "
    "w_lineage_epigenetic=[0.0,0.03,0.05,0.07], seeds=[42,1,99], w_diff_epigenetic=0.0"
)
insert_row += 1

def set_cell(row, col_name, value):
    if col_name in col_idx:
        ws.cell(row, col_idx[col_name]).value = value

for r in runs:
    set_cell(insert_row, "run_id",           r["run_id"])
    set_cell(insert_row, "mode",             FIXED["mode_inh"])
    set_cell(insert_row, "seed",             r["seed"])
    set_cell(insert_row, "t_max",            FIXED["t_max"])
    set_cell(insert_row, "init_hsc",         FIXED["init_hsc"])
    set_cell(insert_row, "cent_stem",        FIXED["cent_stem"])
    set_cell(insert_row, "cent_stress",      FIXED["cent_stress"])
    set_cell(insert_row, "age_cap",          FIXED["age_cap"])
    set_cell(insert_row, "stress_acc",       FIXED["stress_acc"])
    set_cell(insert_row, "stem_drift",       FIXED["stem_drift"])
    set_cell(insert_row, "w_div_stress",     FIXED["w_div_stress"])
    set_cell(insert_row, "w_div_stemness",   FIXED["w_div_stemness"])
    set_cell(insert_row, "w_div_repl",       FIXED["w_div_repl"])
    set_cell(insert_row, "w_diff_stress",    FIXED["w_diff_stress"])
    set_cell(insert_row, "w_diff_stemness",  FIXED["w_diff_stemness"])
    set_cell(insert_row, "w_diff_repl",      FIXED["w_diff_repl"])
    set_cell(insert_row, "w_apo_stress",     FIXED["w_apo_stress"])
    set_cell(insert_row, "w_apo_stemness",   FIXED["w_apo_stemness"])
    set_cell(insert_row, "w_apo_repl",       FIXED["w_apo_repl"])
    set_cell(insert_row, "final_n",          r.get("final_n"))
    set_cell(insert_row, "HSC",              r.get("HSC"))
    set_cell(insert_row, "HSC_%",            r.get("HSC_pct"))
    set_cell(insert_row, "MPP",              r.get("MPP"))
    set_cell(insert_row, "MPP_%",            r.get("MPP_pct"))
    set_cell(insert_row, "events",           r.get("n_events"))
    set_cell(insert_row, "notes",            f"wle={r['wle']}_{r['status']}")
    set_cell(insert_row, "series",           FIXED["series"])
    set_cell(insert_row, "CLP_%",            r.get("CLP_pct"))
    set_cell(insert_row, "CMP_%",            r.get("CMP_pct"))
    set_cell(insert_row, "B_cell_%",         r.get("B_cell_pct"))
    set_cell(insert_row, "T_cell_%",         r.get("T_cell_pct"))
    set_cell(insert_row, "Myeloid_%",        r.get("Myeloid_pct"))
    set_cell(insert_row, "Erythroid_%",      r.get("Erythroid_pct"))
    set_cell(insert_row, "mature_%",         r.get("mature_pct"))
    set_cell(insert_row, "progenitor_%",     r.get("progenitor_pct"))
    set_cell(insert_row, "HSC_to_MPP_ratio", r.get("hsc_mpp_ratio"))
    set_cell(insert_row, "baseline_flag",    "candidate")
    set_cell(insert_row, "hypothesis",       FIXED["hypothesis"])
    insert_row += 1

wb.save(WORKBOOK_PATH)
print(f"Saved. Total rows now: {ws.max_row}")

flags = [
    ws.cell(r, col_idx["baseline_flag"]).value
    for r in range(2, ws.max_row + 1)
    if ws.cell(r, col_idx["baseline_flag"]).value
]
print("baseline_flag summary:", dict(Counter(flags)))
